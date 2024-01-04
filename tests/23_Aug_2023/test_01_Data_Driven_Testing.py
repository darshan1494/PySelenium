'''
Navigate to this URL - https://app.vwo.com/
Try with different Username, Password
admin, admin
admin, admin123
admin, password123
contact+augg@testingacademy.com, Wingify@123
'''
import time

import pytest
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By

# We are going to use  Data Driven Testing framework
# Pytest basically provides how you can perform  Data Driven Testing using parameterized functions

#How to work with Excel?
# Install openpyxl. It basically supports reading the file by using load_workbook

# Read data from the Excel & fetch the data in the form of list/arrays
def get_test_data():
    workbook = load_workbook("test_data.xlsx") # Read from the file
    sheet = workbook.active
    data = [ ]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row) # whatever the username & password is present, it will read the rows
    return data

@pytest.fixture # Fixture is a method in which code can be executed and basically injects whatever it returns to the following test case functions
def driver():
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://app.vwo.com/")
    yield driver # yield can be used only in case of fixtures.
    driver.quit() # Once webdriver is done with its execution, yield helps to close the browser.
    # In Fixture we can Read a File, Read from DB, Create Webdriver, set data, configurations...

#Why it is called Data driven testing framework?
#We have written below single function but not multiple functions which is drived from def get_test_data()
@pytest.mark.parametrize("username, password, result", get_test_data()) # From data.append(row) data is added/injecting to get_test_data(). The moment we do parametrize, it will start reading the file; Whatever the username & password it was able to find in the rows,it will automatically add them into parameters (username,password) & it gets printed.
def test_vwo_login(driver, username,password, result):
    # Navigate to URL
    # Enter correct username & password
    # click
    # Verify if the login is working fine (Yes or No)
    email = driver.find_element(By.ID,"login-username")
    email.send_keys(username)

    pass_word = driver.find_element(By.ID, "login-password")
    pass_word.send_keys(password)

    sign_in=driver.find_element(By.ID, "js-login-btn")
    sign_in.click()

    time.sleep(5)

    print(username,password,driver.current_url)
    if result == "fail":
        error_message = driver.find_element(By.ID, "js-notification-box-msg").text
        assert error_message in "Your email, password, IP address or location did not match"
    else:
        assert "https://app.vwo.com/#/dashboard" in driver.current_url

    # Alternate way to handle above logic if we don't know whether it is pass or fail:
    # print(username, password, driver.current_url)
    # if result == "fail":
    #     failed_tc = driver.find_element(By.ID, "js-notification-box-msg").is_displayed()
    #     if failed_tc:
    #         assert failed_tc in "Your email, password, IP address or location did not match"
    # else:
    #     assert "https://app.vwo.com/#/dashboard" in driver.current_url # If error message is not displayed assertion will be different, which means it's successful login








